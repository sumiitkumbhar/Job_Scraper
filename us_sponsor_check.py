"""
US H-1B sponsorship-history cross-check — custom add-on for Sumit's fork,
the USA-side counterpart to sponsor_check.py's UK Home Office register check.

Runs AFTER the normal scrapers (scrape_jobs.py) have written output/all_jobs.json.

There is no US equivalent of the UK's single official "register of licensed
sponsors" CSV. Instead this script uses the US Department of Labor's public
LCA (Labor Condition Application) Disclosure Data — the same free, no-API-key
government dataset every H-1B-sponsor lookup site (myvisajobs, h1bdata.info,
h1bgrader, etc.) is built on. Every H-1B petition requires a certified LCA
first, so "this employer has certified LCAs on file" is a solid, evidence-
based signal that they sponsor work visas — not a guess or a hand-picked
company list.

Writes two files, matching sponsor_check.py's shape:
  output/us_sponsor_status.json     — one row per unique company seen in this run
  output/jobs_with_us_sponsor.csv   — every job in all_jobs.json + US sponsor columns

Design notes:
  - The DOL performance page is scraped for current "LCA_Disclosure_Data_FY*_Q*.xlsx"
    links because, like the UK register, the exact filenames roll forward every
    quarter with no stable link. QUARTERS_TO_FETCH (default 2) picks the N most
    recent quarters found, so this keeps working as new quarters are published
    without ever needing a code change. If no links are found, the script exits
    cleanly and leaves previous output untouched (same "preserve on failure"
    pattern as the rest of this repo).
  - Each quarterly file is a large (~200-300MB) Excel workbook covering every
    LCA filed that quarter across all US employers, all visa classes (H-1B,
    H-1B1 Chile/Singapore, E-3 Australia). Only rows with VISA_CLASS == "H-1B"
    and a CASE_STATUS of "Certified" or "Certified - Withdrawn" count as a
    sponsorship signal (a certified LCA is DOL sign-off; "withdrawn" after
    certification doesn't undo that the employer was approved to sponsor).
  - Matching is a normalized substring match (strip Inc/LLC/Corp/Ltd/Group/etc,
    lowercase, collapse whitespace) in both directions, with the same short-name
    guard as the UK checker. Same caveat applies: always confirm manually before
    relying on this for a specific application — a scraped "company" is
    sometimes a recruiter, and LCA history doesn't guarantee a company will
    sponsor a *new* hire today.
  - No API key, no paid service, no scope creep into approved-petition counts
    or wage analytics — deliberately kept to the same "confirmed / likely /
    not found" shape as the UK check so the two are easy to read side by side.
  - Needs network access + openpyxl, which is why this runs in GitHub Actions
    (unrestricted egress) rather than locally in a sandboxed environment.

Run manually:
    python us_sponsor_check.py

Wire into CI: see .github/workflows/us_sponsor_watch.yml (runs daily, after
the other watchers, and commits the two output files like every other source
in this repo).
"""

from __future__ import annotations

import csv
import json
import os
import re
import sys
import urllib.error
import urllib.request
from datetime import datetime, timezone

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "output")
ALL_JOBS_PATH = os.path.join(OUTPUT_DIR, "all_jobs.json")
US_SPONSOR_STATUS_PATH = os.path.join(OUTPUT_DIR, "us_sponsor_status.json")
US_JOBS_CSV_PATH = os.path.join(OUTPUT_DIR, "jobs_with_us_sponsor.csv")
TMP_XLSX_PATH = os.path.join(OUTPUT_DIR, "_tmp_lca_disclosure.xlsx")

DOL_PERFORMANCE_PAGE = "https://www.dol.gov/agencies/eta/foreign-labor/performance"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (compatible; JobScraperUSSponsorCheck/1.0; "
        "personal job-search tool; no scraping of dol.gov beyond this one page+files)"
    )
}
REQUEST_TIMEOUT = 60
DOWNLOAD_TIMEOUT = 600  # these files are large (~200-300MB each)
CHUNK_SIZE = 1024 * 1024

# How many of the most recent quarterly LCA disclosure files to combine. Each
# is ~200-300MB, so keep this modest — 2 quarters (~6 months) is a good
# balance of coverage vs. daily-run time/bandwidth. Raise if you want deeper
# history; the discovery logic below always picks the N *most recent*.
QUARTERS_TO_FETCH = 2

ACCEPTED_CASE_STATUSES = {"certified", "certified - withdrawn"}

ORG_SUFFIX_RE = re.compile(
    r"\b(incorporated|inc|llc|l l c|corp|corporation|company|co|llp|pllc|pc|"
    r"ltd|limited|group|holdings?|the)\b\.?",
    re.I,
)
NON_ALNUM_RE = re.compile(r"[^a-z0-9]+")


def normalize_org(name: str) -> str:
    """Lowercase, strip common US legal suffixes and punctuation, collapse whitespace."""
    if not name:
        return ""
    n = name.lower()
    n = ORG_SUFFIX_RE.sub(" ", n)
    n = NON_ALNUM_RE.sub(" ", n)
    return " ".join(n.split())


def fetch_text(url: str, timeout: int = REQUEST_TIMEOUT) -> str:
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read().decode("utf-8", "ignore")


def download_to_file(url: str, dest_path: str, timeout: int = DOWNLOAD_TIMEOUT) -> None:
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=timeout) as resp, open(dest_path, "wb") as out:
        while True:
            chunk = resp.read(CHUNK_SIZE)
            if not chunk:
                break
            out.write(chunk)


def find_lca_file_urls() -> list[str]:
    """Scrape the DOL performance page for LCA_Disclosure_Data_FY*_Q*.xlsx links,
    and return the URLs of the QUARTERS_TO_FETCH most recent ones."""
    try:
        html = fetch_text(DOL_PERFORMANCE_PAGE)
    except (urllib.error.URLError, TimeoutError) as e:
        print(f"  WARNING: could not load DOL performance page: {e}")
        return []

    matches = re.findall(
        r'href="([^"]*LCA_Disclosure_Data_FY(\d{4})_Q(\d)\.xlsx)"',
        html,
        re.I,
    )
    if not matches:
        print("  WARNING: no LCA_Disclosure_Data_FY*_Q*.xlsx links found on the DOL page")
        return []

    # Dedupe by (year, quarter), keep the href, sort newest-first.
    by_period: dict[tuple[int, int], str] = {}
    for href, year, quarter in matches:
        url = href if href.startswith("http") else f"https://www.dol.gov{href}"
        url = url.replace("dol.gov//", "dol.gov/")  # the page has a doubled-slash link
        by_period[(int(year), int(quarter))] = url

    newest_first = sorted(by_period.keys(), reverse=True)
    chosen = newest_first[:QUARTERS_TO_FETCH]
    return [by_period[k] for k in chosen]


def col_index_map(header_row: tuple) -> dict[str, int]:
    """Map UPPER_SNAKE column name -> index, tolerant of stray whitespace."""
    idx = {}
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().upper()
        idx[key] = i
    return idx


def process_workbook(path: str, lookup: dict[str, dict]) -> int:
    """Stream rows from one LCA disclosure workbook into `lookup`, keyed by
    normalized employer name. Returns the number of qualifying rows counted."""
    try:
        import openpyxl
    except ImportError:
        print("  ⛔ openpyxl not installed — add it to requirements.txt.")
        return 0

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    header = next(rows, None)
    if not header:
        wb.close()
        return 0
    idx = col_index_map(header)

    required = ["EMPLOYER_NAME", "VISA_CLASS", "CASE_STATUS"]
    if not all(k in idx for k in required):
        print(f"  WARNING: expected columns not found in {os.path.basename(path)}; saw {list(idx.keys())[:15]}...")
        wb.close()
        return 0

    name_i = idx["EMPLOYER_NAME"]
    visa_i = idx["VISA_CLASS"]
    status_i = idx["CASE_STATUS"]
    state_i = idx.get("WORKSITE_STATE")
    title_i = idx.get("JOB_TITLE")
    soc_i = idx.get("SOC_TITLE")
    decision_i = idx.get("DECISION_DATE")

    counted = 0
    for row in rows:
        try:
            visa_class = (row[visa_i] or "").strip().upper() if row[visa_i] else ""
            case_status = (row[status_i] or "").strip().lower() if row[status_i] else ""
        except IndexError:
            continue
        if visa_class != "H-1B" or case_status not in ACCEPTED_CASE_STATUSES:
            continue

        raw_name = (row[name_i] or "").strip() if name_i < len(row) else ""
        norm = normalize_org(raw_name)
        if len(norm) < 3:
            continue

        entry = lookup.setdefault(norm, {
            "official_name": raw_name,
            "certified_lca_count": 0,
            "states": set(),
            "sample_titles": set(),
            "most_recent_decision_date": "",
        })
        entry["certified_lca_count"] += 1
        if state_i is not None and state_i < len(row) and row[state_i]:
            entry["states"].add(str(row[state_i]).strip())
        if title_i is not None and title_i < len(row) and row[title_i] and len(entry["sample_titles"]) < 5:
            entry["sample_titles"].add(str(row[title_i]).strip())
        elif soc_i is not None and soc_i < len(row) and row[soc_i] and len(entry["sample_titles"]) < 5:
            entry["sample_titles"].add(str(row[soc_i]).strip())
        if decision_i is not None and decision_i < len(row) and row[decision_i]:
            d = str(row[decision_i])
            if d > entry["most_recent_decision_date"]:
                entry["most_recent_decision_date"] = d
        counted += 1

    wb.close()
    return counted


def match_company(company: str, lookup: dict[str, dict]) -> dict:
    norm = normalize_org(company)
    if not norm:
        return {"status": "unknown", "matched_name": None, "certified_lca_count": 0}

    if norm in lookup:
        e = lookup[norm]
        return {
            "status": "confirmed",
            "matched_name": e["official_name"],
            "certified_lca_count": e["certified_lca_count"],
        }

    if len(norm) >= 4:
        for key, e in lookup.items():
            if len(key) < 4:
                continue
            if norm in key or key in norm:
                return {
                    "status": "likely",
                    "matched_name": e["official_name"],
                    "certified_lca_count": e["certified_lca_count"],
                }

    return {"status": "not_found", "matched_name": None, "certified_lca_count": 0}


def main() -> int:
    if not os.path.exists(ALL_JOBS_PATH):
        print(f"  No {ALL_JOBS_PATH} yet — run scrape_jobs.py first. Skipping.")
        return 0

    with open(ALL_JOBS_PATH, encoding="utf-8") as f:
        all_jobs_doc = json.load(f)
    jobs = all_jobs_doc.get("jobs", [])
    if not jobs:
        print("  all_jobs.json has no jobs yet. Skipping.")
        return 0

    print("🔎 Checking US H-1B LCA disclosure data...")
    file_urls = find_lca_file_urls()
    if not file_urls:
        print("  ⛔ Could not locate any LCA disclosure files this run; leaving previous US sponsor output untouched.")
        return 0
    print(f"  Using {len(file_urls)} most recent quarterly file(s):")
    for u in file_urls:
        print(f"    {u}")

    lookup: dict[str, dict] = {}
    total_rows = 0
    for url in file_urls:
        try:
            print(f"  Downloading {url} ...")
            download_to_file(url, TMP_XLSX_PATH)
            n = process_workbook(TMP_XLSX_PATH, lookup)
            print(f"    {n} certified H-1B rows counted from this file.")
            total_rows += n
        except (urllib.error.URLError, TimeoutError, OSError) as e:
            print(f"  WARNING: could not process {url}: {e}")
        finally:
            if os.path.exists(TMP_XLSX_PATH):
                os.remove(TMP_XLSX_PATH)

    if not lookup:
        print("  ⛔ No usable LCA data parsed this run; leaving previous US sponsor output untouched.")
        return 0
    print(f"  Loaded {len(lookup)} unique employers with certified H-1B LCAs ({total_rows} qualifying rows total).")

    company_status: dict[str, dict] = {}
    enriched_rows = []
    for job in jobs:
        company = job.get("company", "") or ""
        if company not in company_status:
            company_status[company] = match_company(company, lookup)
        status = company_status[company]

        enriched_rows.append({
            "title": job.get("title", ""),
            "company": company,
            "location": job.get("location", ""),
            "us_sponsor_status": status["status"],
            "us_sponsor_matched_name": status["matched_name"] or "",
            "certified_lca_count": status["certified_lca_count"],
            "salary": job.get("salary", ""),
            "date_posted": job.get("date_posted", ""),
            "ats": job.get("ats", ""),
            "url": job.get("url", ""),
        })

    counts = {"confirmed": 0, "likely": 0, "not_found": 0, "unknown": 0}
    for s in company_status.values():
        counts[s["status"]] = counts.get(s["status"], 0) + 1

    # JSON can't serialize sets — convert before dumping.
    serializable_companies = {}
    for company, status in company_status.items():
        serializable_companies[company] = status

    with open(US_SPONSOR_STATUS_PATH, "w", encoding="utf-8") as f:
        json.dump({
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "source_files": file_urls,
            "employer_count": len(lookup),
            "company_counts": counts,
            "companies": serializable_companies,
        }, f, indent=2)

    with open(US_JOBS_CSV_PATH, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(enriched_rows[0].keys()) if enriched_rows else [])
        writer.writeheader()
        writer.writerows(enriched_rows)

    print(
        f"  ✅ {len(company_status)} companies checked — "
        f"{counts.get('confirmed', 0)} confirmed sponsors, "
        f"{counts.get('likely', 0)} likely matches, "
        f"{counts.get('not_found', 0)} not found."
    )
    print(f"  📄 Saved {US_SPONSOR_STATUS_PATH} and {US_JOBS_CSV_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
